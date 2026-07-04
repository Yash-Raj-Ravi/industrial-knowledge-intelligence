from openpyxl import load_workbook

def parse_xlsx(file_path: str) -> str:
    workbook = load_workbook(file_path,data_only=True)
    sheets_text=[]
    for sheet in workbook:
        for row in sheet.iter_rows(values_only = True):
            text = ",".join(str(cell) for cell in row if cell is not None) # Converting each cell to string
            if text.strip():
                sheets_text.append(text)

    return "\n".join(sheets_text)